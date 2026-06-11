"""
Generate the reader-facing Supplementary Data Excel workbook from compostable-packaging
result-calculation outputs. This script is the presentation layer: it selects
reader-facing sheets, masks paths, filters codebook sheets, and applies
workbook formatting without changing analytical results.

Design choices
--------------
- 00_Result_Summary is the main reporting sheet and contains default/all-source
  share calculations in a reasoning sequence.
- Source_Level_Coding and Application_Decisions are kept clean: no regex/matched-text
  trace columns in reader-facing sheets.
- Rule traceability remains in audit CSV files, while the codebook sheets show
  representative include and exclude regex examples for each category.
- No Excel structured Table XML objects are created, avoiding repair warnings.
- No separate validation-check worksheet is generated in the reader-facing workbook.
"""
from __future__ import annotations

import argparse
import datetime as dt
import re
import zipfile
from pathlib import Path
from typing import Dict, Optional, Tuple

import pandas as pd
from pandas.api.types import is_object_dtype, is_string_dtype
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent if "__file__" in globals() else Path.cwd()
DEFAULT_INPUT = SCRIPT_DIR / "results"
DEFAULT_OUTPUT_XLSX = SCRIPT_DIR / "Supplementary_Data.xlsx"

REQUIRED_FILES = {
    "reference_coding": "source_level_coding_display.csv",
    "reference_shares": "source_level_category_shares_display.csv",
    "rejection_rationale": "rejection_rationale_reporting_shares_display.csv",
    "application_long": "application_decision_long_display.csv",
    "treatment_route_audit_summary": "treatment_route_not_stated_audit_summary_display.csv",
    "application_shares": "application_decision_shares_display.csv",
    "application_compatibility": "application_compatibility_scenarios_display.csv",
    "certification_sufficiency_summary": "certification_sufficiency_summary_display.csv",
    "certification_sufficiency_by_application": "certification_sufficiency_by_application_display.csv",
    "certification_sufficiency_bubble_matrix": "certification_sufficiency_bubble_matrix_display.csv",
    "certification_sufficiency_by_certification_basis": "certification_sufficiency_by_certification_basis_display.csv",
    "certification_sufficiency_by_source_authority": "certification_sufficiency_by_source_authority_display.csv",
    "certification_sufficiency_long": "certification_sufficiency_long_display.csv",
    "source_level_sensitivity": "source_level_sensitivity_display.csv",
    "codebook": "codebook_display.csv",
}

SUMMARY_INDICATOR_ORDER = [
    "Source Authority",
    "Stated Organic-waste Treatment Route",
    "Acceptance",
    "Rejection Rationale",
]

SUMMARY_SECTION_TITLES = {
    "Source Authority": "1.1 Source Authority",
    "Stated Organic-waste Treatment Route": "1.2 Stated Organic-waste Treatment Route",
    "Acceptance": "1.3 Acceptance",
    "Rejection Rationale": "1.4 Rejection Rationale",
    "Certification and Approval Basis": "3.1 Certification and Approval Basis",
}

REJECTION_RATIONALE_UNKNOWN_DISPLAY = "Reason not explicitly stated"
REJECTION_RATIONALE_NOT_APPLICABLE = "Not applicable - not a rejection/restriction case"
REJECTION_RATIONALE_REPORTED_CATEGORIES = {
    "Explicit no-packaging / positive-list restriction",
    "Compost/digestate quality or contamination concern",
    "No compatible organic-treatment route",
    "Slow degradation / residence-time mismatch",
    "AD/biogas incompatibility",
    "Pre-treatment/screening/equipment constraint",
}

CODEBOOK_SHEET_NAMES = {
    "Source Authority": "CB_Source_Authority",
    "Stated Organic-waste Treatment Route": "CB_Treatment_Route",
    "Acceptance": "CB_Acceptance",
    "Rejection Rationale": "CB_Rejection_Rationale",
    "Application Group": "CB_Application_Group",
    "Application Decision": "CB_Application_Decision",
    "Certification and Approval Basis": "CB_Certification_Basis",
    "Certification Sufficiency": "CB_Certification_Suff",
    "Source-level Sensitivity Group": "CB_Source_Sensitivity",
}
# Canonical reader-facing certification category labels.
# This allows the Excel generator to clean older display CSV outputs produced
# before the category name was unified.
CERTIFICATION_CATEGORY_RENAME = {
    "BNQ / AS 4736 / DINplus / national standards": "Country-specific standards/certifications (BNQ / AS 4736 / DINplus)",
    "National standards (BNQ / AS 4736 / DINplus etc.)": "Country-specific standards/certifications (BNQ / AS 4736 / DINplus)",
    "National standards": "Country-specific standards/certifications (BNQ / AS 4736 / DINplus)",
    "Country-specific standards/certifications (BNQ / AS 4736 / DINplus etc.)": "Country-specific standards/certifications (BNQ / AS 4736 / DINplus)",
    "Country-specific standards/certifications (BNQ / AS 4736 / DINplus)": "Country-specific standards/certifications (BNQ / AS 4736 / DINplus)",
}




# Backward-compatible reader-facing labels for the revised Certification Sufficiency taxonomy.
# The current denominator is certification-relevant source × application-group
# records only. Generic compostable/biodegradable wording and no-standard cases
# are intentionally outside Certification Sufficiency.
CERTIFICATION_SUFFICIENCY_RENAME = {
    "Named certification/approval supports acceptance": "Certification accepted",
    "Certification sufficient for ordinary acceptance": "Certification accepted",
    "Certification/approval required but not sufficient": "Certification conditionally accepted",
    "Rejected despite named certification/approval": "Certification rejected",
    "Certification mentioned but operationally rejected": "Certification rejected",
    "Certification irrelevant because application group is excluded": "Certification rejected",
    "Acceptance unclear or not stated": "Certification unclear or not stated",
    "Compatibility unclear or not stated": "Certification unclear or not stated",
}

CERTIFICATION_SUFFICIENCY_COLUMNS_ORDER = [
    "Certification accepted",
    "Certification conditionally accepted",
    "Certification rejected",
    "Certification unclear or not stated",
]



FILL_TITLE = PatternFill("solid", fgColor="0F766E")
FILL_SECTION = PatternFill("solid", fgColor="CCFBF1")
FILL_HEADER = PatternFill("solid", fgColor="D9EAD3")
FILL_NOTE = PatternFill("solid", fgColor="FEF3C7")
FILL_LIGHT = PatternFill("solid", fgColor="F8FAFC")
FONT_TITLE = Font(bold=True, color="FFFFFF", size=15)
FONT_SECTION = Font(bold=True, color="134E4A", size=12)
FONT_HEADER = Font(bold=True, color="1F2937")
FONT_SMALL = Font(size=9, color="475569")
SIDE = Side(style="thin", color="CBD5E1")
BORDER = Border(left=SIDE, right=SIDE, top=SIDE, bottom=SIDE)


def resolve_path(path_like: Optional[str | Path]) -> Optional[Path]:
    if path_like is None:
        return None
    p = Path(path_like).expanduser()
    if not p.is_absolute():
        p = SCRIPT_DIR / p
    return p.resolve()


def read_display_tables(input_path: Path) -> Tuple[Dict[str, pd.DataFrame], str]:
    if input_path.is_dir():
        display_dir = input_path / "display" if (input_path / "display").exists() else input_path
        out: Dict[str, pd.DataFrame] = {}
        for key, filename in REQUIRED_FILES.items():
            path = display_dir / filename
            if not path.exists():
                legacy = filename.replace("source_level_coding", "reference_level_coding").replace("source_level_category_shares", "reference_level_category_shares")
                legacy_path = display_dir / legacy
                if legacy_path.exists():
                    path = legacy_path
                else:
                    raise FileNotFoundError(f"Missing required display CSV: {path}")
            out[key] = pd.read_csv(path)
        return out, f"display folder: {display_dir}"
    if input_path.suffix.lower() == ".zip":
        out = {}
        with zipfile.ZipFile(input_path, "r") as zf:
            names = set(zf.namelist())
            for key, filename in REQUIRED_FILES.items():
                candidates = [n for n in names if Path(n).name == filename]
                if not candidates:
                    legacy = filename.replace("source_level_coding", "reference_level_coding").replace("source_level_category_shares", "reference_level_category_shares")
                    candidates = [n for n in names if Path(n).name == legacy]
                if not candidates:
                    raise FileNotFoundError(f"Missing {filename} inside {input_path}")
                with zf.open(candidates[0]) as f:
                    out[key] = pd.read_csv(f)
        return out, f"display zip: {input_path}"
    raise FileNotFoundError(f"Input is neither a display folder nor ZIP: {input_path}")


def sanitize_path_for_workbook(text: str) -> str:
    """Hide user-/institution-specific middle path components in workbook metadata."""
    if not text:
        return text
    prefix = ""
    path_text = str(text)
    if ": " in path_text:
        prefix, path_text = path_text.split(": ", 1)
        prefix = prefix + ": "
    norm_path = path_text.replace("\\", "/")
    parts = [part for part in norm_path.split("/") if part]
    if len(parts) <= 4:
        return prefix + norm_path
    # Keep the drive/root cue and the project tail, but hide personal and institutional folders.
    if parts[0].endswith(":"):
        keep_start = parts[:2]
    else:
        keep_start = parts[:1]
    keep_tail = parts[-3:]
    return prefix + "/".join(keep_start + ["..."] + keep_tail)




def harmonize_certification_category_labels(df: pd.DataFrame) -> pd.DataFrame:
    """Apply canonical Certification and Approval Basis category labels everywhere.

    This cleans both category cells and explanatory codebook text from older
    display CSVs generated before the category name was unified.
    """
    if df is None or df.empty:
        return df

    out = df.copy()

    # Exact replacement for category-like columns.
    for col in out.columns:
        col_l = str(col).lower()
        if (
            col in {
                "category",
                "Certification and Approval Basis",
                "Certification or Approval Basis",
                "certification_basis",
            }
            or "certification" in col_l
        ):
            out[col] = out[col].replace(CERTIFICATION_CATEGORY_RENAME)

    if {"indicator", "category"}.issubset(out.columns):
        mask = out["indicator"].eq("Certification and Approval Basis")
        out.loc[mask, "category"] = out.loc[mask, "category"].replace(CERTIFICATION_CATEGORY_RENAME)

    # Substring replacement for reader-facing codebook text and notes.
    replacements = {
        "BNQ / AS 4736 / DINplus / national standards": "Country-specific standards/certifications (BNQ / AS 4736 / DINplus)",
        "National standards (BNQ / AS 4736 / DINplus etc.)": "Country-specific standards/certifications (BNQ / AS 4736 / DINplus)",
        "National standards (BNQ / AS 4736 / DINplus etc.)": "Country-specific standards/certifications (BNQ / AS 4736 / DINplus)",
        "national/regional standard cue": "country-specific standard/certification cue",
        "national or regional compostability standards": "country-specific compostability standards/certifications",
        "national compostability standard": "country-specific compostability standard/certification",
    }
    for col in [c for c in out.columns if is_object_dtype(out[c].dtype) or is_string_dtype(out[c].dtype)]:
        s = out[col].astype(str)
        for old, new in replacements.items():
            s = s.str.replace(old, new, regex=False)
        out[col] = s

    return out



def harmonize_certification_sufficiency_labels(df: pd.DataFrame) -> pd.DataFrame:
    """Apply canonical Certification Sufficiency category labels in all tables."""
    if df is None or df.empty:
        return df
    out = df.copy()

    # Exact replacement in likely category columns and all object columns.
    for col in [c for c in out.columns if is_object_dtype(out[c].dtype) or is_string_dtype(out[c].dtype)]:
        out[col] = out[col].replace(CERTIFICATION_SUFFICIENCY_RENAME)
        s = out[col].astype(str)
        for old, new in CERTIFICATION_SUFFICIENCY_RENAME.items():
            s = s.str.replace(old, new, regex=False)
        out[col] = s

    # Rename wide cross-tab columns and share columns.
    rename_cols = {}
    for col in out.columns:
        if col in CERTIFICATION_SUFFICIENCY_RENAME:
            rename_cols[col] = CERTIFICATION_SUFFICIENCY_RENAME[col]
        if isinstance(col, str) and col.startswith("share__"):
            suffix = col.replace("share__", "", 1)
            if suffix in CERTIFICATION_SUFFICIENCY_RENAME:
                rename_cols[col] = "share__" + CERTIFICATION_SUFFICIENCY_RENAME[suffix]
    if rename_cols:
        out = out.rename(columns=rename_cols)

    return out

def align_table_terminology(tables: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    """Align CSV/workbook labels with the final manuscript terminology."""
    out = {k: v.copy() for k, v in tables.items()}
    indicator_rename = {
        "Certification and Approval Basis": "Certification and Approval Basis",
        "Certification Sufficiency / Accountability Gap": "Certification Sufficiency",
        "Binary Compatibility Scenario": "Application-level Compatibility Scenario",
        "Application Group": "Application Group",
        "Source-level sensitivity group": "Source-level Sensitivity Group",
    }
    column_rename = {
        "Certification and Approval Basis": "Certification and Approval Basis",
        "Source-level sensitivity group": "Source-level Sensitivity Group",
    }
    drop_codebook_indicators = {"Evidence Role", "Acceptance Condition"}
    for key, df in out.items():
        if df is None or df.empty:
            continue
        df.rename(columns=column_rename, inplace=True)
        if "indicator" in df.columns:
            df["indicator"] = df["indicator"].replace(indicator_rename)
            if key == "codebook":
                df.drop(df[df["indicator"].isin(drop_codebook_indicators)].index, inplace=True)
        if "denominator_basis" in df.columns:
            df["denominator_basis"] = df["denominator_basis"].astype(str).str.replace("reference × detected application group", "source × application group", regex=False)
            df["denominator_basis"] = df["denominator_basis"].astype(str).str.replace("Reference × detected application group", "Source × application group", regex=False)
            df["denominator_basis"] = df["denominator_basis"].astype(str).str.replace("references", "sources", regex=False)
            df["denominator_basis"] = df["denominator_basis"].astype(str).str.replace("References", "Sources", regex=False)
        # Clean explanatory text so removed internal fields do not reappear as headline terminology.
        for col in [c for c in df.columns if is_object_dtype(df[c].dtype) or is_string_dtype(df[c].dtype)]:
            df[col] = (df[col].astype(str)
                .str.replace("Certification or Approval Basis", "Certification and Approval Basis", regex=False)
                .str.replace("Certification/Approval Basis", "Certification and Approval Basis", regex=False)
                .str.replace("Acceptance Condition", "conditional-acceptance cue", regex=False)
                .str.replace("Acceptance + conditional-acceptance cue", "Acceptance + conditional-acceptance cue", regex=False)
                .str.replace("Certification Sufficiency / Accountability Gap", "Certification Sufficiency", regex=False)
                .str.replace("certification-sufficiency", "certification-sufficiency", regex=False)
                .str.replace("certification sufficiency", "certification sufficiency", regex=False)
                .str.replace("Reference × detected application group", "Source × application group", regex=False)
                .str.replace("reference × detected application group", "source × application group", regex=False))
        out[key] = harmonize_certification_sufficiency_labels(harmonize_certification_category_labels(df))
    return out


def clean_value(value):
    if pd.isna(value):
        return None
    return value




def augment_rejection_rationale_codebook(tables: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    """Add reader-facing reporting-status notes to the rejection-rationale codebook.

    The coding script already writes display-ready category labels. The Excel
    generator therefore does not relabel categories; it only adds two workbook
    columns explaining which rejection-rationale categories enter the main
    denominator and which are retained as audit statuses.
    """
    out = {k: v.copy() for k, v in tables.items()}
    if "codebook" in out and {"indicator", "category"}.issubset(out["codebook"].columns):
        cb = out["codebook"].copy()
        mask = cb["indicator"].eq("Rejection Rationale")
        cb.loc[mask, "main_share_denominator_status"] = cb.loc[mask, "category"].map(rejection_rationale_reporting_status)
        cb.loc[mask, "reporting_note"] = cb.loc[mask, "category"].map(rejection_rationale_reporting_note)
        out["codebook"] = cb
    return out

def rejection_rationale_reporting_status(category: str) -> str:
    if category in REJECTION_RATIONALE_REPORTED_CATEGORIES:
        return "Included in main share"
    if category == REJECTION_RATIONALE_UNKNOWN_DISPLAY:
        return "Excluded from main share; audit/count only"
    if category == REJECTION_RATIONALE_NOT_APPLICABLE:
        return "Excluded from main share; outside rejection-rationale denominator"
    return "Check category status"


def rejection_rationale_reporting_note(category: str) -> str:
    if category in REJECTION_RATIONALE_REPORTED_CATEGORIES:
        return "Substantive stated rationale among rejected/restricted cases."
    if category == REJECTION_RATIONALE_UNKNOWN_DISPLAY:
        return "The source indicates rejection/restriction, but the rationale is not explicitly stated; retained for audit but not used as a rationale category."
    if category == REJECTION_RATIONALE_NOT_APPLICABLE:
        return "No rejection/restriction rationale is applicable for this source; retained to avoid blank coding values."
    return "Category is not part of the predefined rejection-rationale reporting set."


def rejection_rationale_audit_counts(ref: pd.DataFrame) -> pd.DataFrame:
    if "Rejection Rationale" not in ref.columns:
        return pd.DataFrame(columns=["Audit status", "n", "Reporting treatment"])
    counts = ref["Rejection Rationale"].value_counts(dropna=False).to_dict()
    return pd.DataFrame([
        [REJECTION_RATIONALE_UNKNOWN_DISPLAY, int(counts.get(REJECTION_RATIONALE_UNKNOWN_DISPLAY, 0)), "Retained in source-level coding and codebook, excluded from main stated-rationale share."],
        [REJECTION_RATIONALE_NOT_APPLICABLE, int(counts.get(REJECTION_RATIONALE_NOT_APPLICABLE, 0)), "Retained in source-level coding and codebook, excluded because no rejection/restriction rationale is applicable."],
    ], columns=["Audit status", "n", "Reporting treatment"])

def safe_sheet_name(name: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "_", str(name))[:31] or "Sheet"


def add_sheet_title(ws, title: str, subtitle: str = "", colspan: int = 8) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=colspan)
    cell = ws.cell(1, 1, title)
    cell.fill = FILL_TITLE
    cell.font = FONT_TITLE
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=colspan)
        c = ws.cell(2, 1, subtitle)
        c.fill = FILL_NOTE
        c.font = FONT_SMALL
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[2].height = 46
        return 4
    return 3


def add_section(ws, row: int, title: str, question: str = "", colspan: int = 8) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=colspan)
    c = ws.cell(row, 1, title)
    c.fill = FILL_SECTION
    c.font = FONT_SECTION
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1
    if question:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=colspan)
        q = ws.cell(row, 1, f"Question answered: {question}")
        q.fill = FILL_LIGHT
        q.font = FONT_SMALL
        q.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1
    return row


def write_dataframe(ws, df: pd.DataFrame, start_row: int, start_col: int = 1, header: bool = True) -> Tuple[int, int]:
    r = start_row
    if header:
        for c, h in enumerate(df.columns, start=start_col):
            ws.cell(r, c, h)
        r += 1
    for row in df.itertuples(index=False, name=None):
        for c, value in enumerate(row, start=start_col):
            ws.cell(r, c, clean_value(value))
        r += 1
    return r - start_row, len(df.columns)


def style_range(ws, start_row: int, end_row: int, start_col: int, end_col: int, header_row: Optional[int] = None) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if header_row:
        for row in ws.iter_rows(min_row=header_row, max_row=header_row, min_col=start_col, max_col=end_col):
            for c in row:
                c.fill = FILL_HEADER
                c.font = FONT_HEADER
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def format_numeric_columns(ws, df: pd.DataFrame, start_row: int, header: bool = True) -> None:
    data_start = start_row + (1 if header else 0)
    for idx, col in enumerate(df.columns, start=1):
        name = str(col).lower()
        if "percent" in name:
            fmt = "0.0"
        elif "share" in name:
            fmt = "0.000"
        elif name in {"n", "denominator", "value", "evidence_records"} or name.startswith("n_") or "denominator" in name:
            fmt = "0"
        else:
            continue
        for r in range(data_start, data_start + len(df)):
            ws.cell(r, idx).number_format = fmt


def autofit_reasonable(ws, df: pd.DataFrame, start_col: int = 1) -> None:
    """Fit worksheet dimensions to content while keeping the workbook readable."""
    if df is None or df.empty:
        return

    for j, col in enumerate(df.columns, start=start_col):
        col_letter = get_column_letter(j)
        name = str(col).lower()
        values = [str(col)] + ["" if pd.isna(x) else str(x) for x in df[col].tolist()]
        lengths = sorted([len(v) for v in values] or [10])
        max_len = max(lengths + [10])
        q95_len = lengths[int(0.95 * (len(lengths) - 1))] if len(lengths) > 1 else max_len
        base_len = max(q95_len, min(max_len, 80))

        if "url" in name:
            width = min(max(base_len * 0.9, 34), 70)
        elif any(k in name for k in ["definition", "examples", "note", "question", "basis", "rule", "description", "interpretation", "rationale", "breakdown"]):
            width = min(max(base_len * 0.75, 28), 65)
        elif any(k in name for k in ["title", "source_label"]):
            width = min(max(base_len * 0.75, 26), 55)
        elif any(k in name for k in ["category", "indicator", "application", "acceptance", "certification", "decision"]):
            width = min(max(base_len * 0.85, 20), 44)
        else:
            width = min(max(base_len + 2, 10), 34)
        ws.column_dimensions[col_letter].width = round(width, 1)

    max_col = min(ws.max_column, max(len(df.columns), 1))
    for row_idx in range(1, min(ws.max_row, 1200) + 1):
        max_lines = 1
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            text = "" if cell.value is None else str(cell.value)
            width = ws.column_dimensions[get_column_letter(col_idx)].width or 18
            manual_lines = text.count("\n") + 1
            wrapped_lines = max(1, int(len(text) / max(width * 1.2, 12)) + 1)
            max_lines = max(max_lines, manual_lines, wrapped_lines)
        height = 24 if row_idx <= 3 else min(max(18, max_lines * 15), 120)
        ws.row_dimensions[row_idx].height = height


def create_readme(wb: Workbook, source_desc: str) -> None:
    ws = wb.create_sheet("00_Read_Me", 0)
    ws.sheet_view.showGridLines = False
    row = add_sheet_title(ws, "Compostable-packaging compatibility evidence workbook", "Supplementary Data workbook generated from the reproducible result-calculation workflow.", 8)
    df = pd.DataFrame([
        ["Generated", dt.datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Input", "Generated display CSVs from results/display, produced by result_calculation.py from web_sources.zip."],
        ["Main result logic", "00_Result_Summary is organized by three analysis routes: (1) source-level operational rules, (2) application-level compatibility scenario output, and (3) certification sufficiency for certification-relevant records."],
        ["Rejection-rationale denominator", "The main Rejection Rationale share is based only on rejected/restricted sources with an explicitly stated rationale. Reason not explicitly stated and Not applicable are retained as audit statuses, but are excluded from that main share."],
        ["Audit traceability", "Regex, rule, and matched-text traceability is saved in the audit CSV files generated by the coding script."],
        ["Source-level sensitivity", "Source-level Sensitivity Group compares all included sources, national/supranational sources, and sources below national level. It is the workbook sensitivity check, separate from the application-level scenario assumptions."],
    ], columns=["Item", "Description"])
    nrows, ncols = write_dataframe(ws, df, row)
    style_range(ws, row, row + nrows - 1, 1, ncols, row)
    format_numeric_columns(ws, df, row)
    row += nrows + 2

    contents = pd.DataFrame([
        ["00_Result_Summary", "Main result tables organized by the three analysis routes."],
        ["01_Source_Level_Coding", "Clean source-level coding sheet. After source metadata, only source-level indicators and the source-level sensitivity grouping are shown."],
        ["02_Application_Records", "Application-level records used to build the application-level compatibility scenario output. Application Group and Application Decision are the two coded dimensions."],
        ["03_Source_Level_Sensitivity", "Sensitivity check by source level: all included sources, national/supranational sources, and sources below national level."],
        ["04_Country_Coverage", "Descriptive country/region coverage and source counts."],
        ["05_Application_by_Country", "Country-grouped application-level view. Columns A–E are manuscript-ready: source count, covered jurisdictions, observed compatible application groups under the central scenario, and observed rejected application groups."],
        ["06_Certification_Suff", "Dedicated certification-sufficiency sheet including the source-level certification basis distribution, the four-category certification-sufficiency summary, and wide/long bubble-chart input tables."],
        ["CB_*", "Codebook sheets defining indicators, categories, examples, regex/rule examples, and mapping logic."],
    ], columns=["Sheet", "Purpose"])
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = ws.cell(row, 1, "Workbook contents")
    c.fill = FILL_SECTION
    c.font = FONT_SECTION
    c.alignment = Alignment(vertical="center")
    row += 1
    nrows, ncols = write_dataframe(ws, contents, row)
    style_range(ws, row, row + nrows - 1, 1, ncols, row)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 110
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 30



def order_certification_sufficiency_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Order Certification Sufficiency rows by analytical mapping, not frequency."""
    if df is None or df.empty or "certification_sufficiency_status" not in df.columns:
        return df
    out = df.copy()
    order_map = {cat: i for i, cat in enumerate(CERTIFICATION_SUFFICIENCY_COLUMNS_ORDER)}
    out["_order"] = out["certification_sufficiency_status"].map(order_map).fillna(999)
    return out.sort_values(["_order", "certification_sufficiency_status"]).drop(columns=["_order"])

def create_result_summary(wb: Workbook, tables: Dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("00_Result_Summary", 1)
    ws.sheet_view.showGridLines = False
    row = add_sheet_title(
        ws,
        "Main results by analysis route",
        "This sheet aligns the workbook with three analysis routes: (1) structured evidence synthesis of local rules, (2) application-level compatibility scenarios, and (3) certification-sufficiency analysis. Rejection-rationale shares use only rejected/restricted sources with an explicitly stated rationale.",
        10,
    )

    ref = tables["reference_coding"]
    app = tables["application_long"]
    compat = tables["application_compatibility"]
    shares = tables["reference_shares"]
    cert_summary = tables.get("certification_sufficiency_summary", pd.DataFrame())
    cert_by_app = tables.get("certification_sufficiency_by_application", pd.DataFrame())
    cert_by_cert = tables.get("certification_sufficiency_by_certification_basis", pd.DataFrame())

    row = add_section(ws, row, "0. Key denominators", "What are the units of analysis behind the reported shares?", 10)
    kpis = pd.DataFrame([
        ["Included sources", len(ref), "Source-level denominator for operational-rule indicators."],
        ["Application records", len(app), "Source × application-group records before scenario conversion."],
        ["Certification sufficiency records", len(tables.get("certification_sufficiency_long", pd.DataFrame())), "Certification-relevant source × application-group records: sources whose Certification and Approval Basis is a certification, standard, or official approval category."],
        ["Sources allocated to sensitivity", int(ref["Source-level Sensitivity Group"].isin({"National/supranational sources", "Sources below national level"}).sum()) if "Source-level Sensitivity Group" in ref.columns else None, "All included sources are allocated to national/supranational sources or sources below national level."],
    ], columns=["Metric", "Value", "Interpretation"])
    nrows, ncols = write_dataframe(ws, kpis, row)
    style_range(ws, row, row + nrows - 1, 1, ncols, row)
    format_numeric_columns(ws, kpis, row)
    row += nrows + 2

    row = add_section(ws, row, "Analysis route 1. Source-level operational rules", "How do organic-waste systems classify compostable packaging at the source level?", 10)
    route1_note = pd.DataFrame([
        ["Purpose", "Summarize the source-level indicators used to describe operational rules: Source Authority, Stated Organic-waste Treatment Route, Acceptance, and Rejection Rationale."],
        ["Unit of analysis", "Included source, except Rejection Rationale, which uses the subset of rejected/restricted sources with an explicitly stated rationale."],
    ], columns=["Item", "Description"])
    nrows, ncols = write_dataframe(ws, route1_note, row)
    style_range(ws, row, row + nrows - 1, 1, ncols, row)
    row += nrows + 2

    for indicator in SUMMARY_INDICATOR_ORDER:
        if indicator == "Rejection Rationale" and "rejection_rationale" in tables:
            sdf = tables["rejection_rationale"].copy()
        else:
            sdf = shares[shares["indicator"] == indicator].copy()
        if sdf.empty:
            continue
        question = sdf["question_answered"].dropna().iloc[0] if "question_answered" in sdf.columns and not sdf["question_answered"].dropna().empty else ""
        row = add_section(ws, row, SUMMARY_SECTION_TITLES.get(indicator, indicator), question, 10)
        show_cols = ["category", "n", "denominator", "share", "share_percent", "denominator_basis"]
        show = sdf[[c for c in show_cols if c in sdf.columns]].copy().sort_values(["n", "category"], ascending=[False, True])
        nrows, ncols = write_dataframe(ws, show, row)
        style_range(ws, row, row + nrows - 1, 1, ncols, row)
        format_numeric_columns(ws, show, row)
        row += nrows + 2

        if indicator == "Stated Organic-waste Treatment Route":
            treatment_summary = tables.get("treatment_route_audit_summary", pd.DataFrame())
            if treatment_summary is not None and not treatment_summary.empty:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
                note = ws.cell(row, 1, "Why treatment route is not stated")
                note.fill = FILL_NOTE
                note.font = FONT_HEADER
                note.alignment = Alignment(wrap_text=True, vertical="top")
                row += 1
                show_audit = treatment_summary[[c for c in ["Treatment Route Coding Note", "n", "denominator", "share", "share_percent", "denominator_basis"] if c in treatment_summary.columns]].copy()
                show_audit = show_audit.rename(columns={"Treatment Route Coding Note": "Route-statement audit category"})
                n_note_rows, n_note_cols = write_dataframe(ws, show_audit, row)
                style_range(ws, row, row + n_note_rows - 1, 1, n_note_cols, row)
                format_numeric_columns(ws, show_audit, row)
                row += n_note_rows + 2

        if indicator == "Rejection Rationale":
            audit = rejection_rationale_audit_counts(ref)
            if not audit.empty:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
                note = ws.cell(row, 1, "Audit statuses excluded from the main rejection-rationale share")
                note.fill = FILL_NOTE
                note.font = FONT_HEADER
                note.alignment = Alignment(wrap_text=True, vertical="top")
                row += 1
                n_audit_rows, n_audit_cols = write_dataframe(ws, audit, row)
                style_range(ws, row, row + n_audit_rows - 1, 1, n_audit_cols, row)
                format_numeric_columns(ws, audit, row)
                row += n_audit_rows + 2

    row = add_section(ws, row, "Analysis route 2. Application-level compatibility scenario output", "How does compatibility differ across compostable-packaging application groups under lower-bound, central, and upper-bound assumptions?", 10)
    route2_note = pd.DataFrame([
        ["What this table is", "A tabulated analytical output, not a single categorical indicator. Rows are application groups. Count columns show the coded application decisions. Scenario columns convert those decisions into compatibility shares."],
        ["Included sources", f"{len(ref)} sources", "Source-level corpus represented before decomposition into source × application-group records."],
        ["Unit of analysis", f"{len(app)} source × application-group records", "One source can contribute multiple application records when it discusses multiple compostable-packaging application groups."],
        ["Scenario conversion", "Lower-bound counts only accepted records as compatible; central counts accepted plus accepted-with-conditions as compatible; upper also treats unclear or not stated records as potentially compatible. Rejected records remain incompatible."],
    ], columns=["Item", "Value", "Description"])
    nrows, ncols = write_dataframe(ws, route2_note, row)
    style_range(ws, row, row + nrows - 1, 1, ncols, row)
    row += nrows + 2
    compat_cols = [
        "application_type", "evidence_records", "n_accepted", "n_accepted_with_conditions", "n_rejected", "n_unclear_or_not_stated",
        "lower_bound_compatible_share", "central_compatible_share", "upper_bound_compatible_share", "denominator_basis",
    ]
    show = compat[[c for c in compat_cols if c in compat.columns]].copy()
    nrows, ncols = write_dataframe(ws, show, row)
    style_range(ws, row, row + nrows - 1, 1, ncols, row)
    format_numeric_columns(ws, show, row)
    row += nrows + 2

    row = add_section(ws, row, "Analysis route 3. Certification sufficiency", "Among certification-relevant records, does certification correspond to acceptance, conditional acceptance, rejection, or uncertainty?", 10)
    route3_note = pd.DataFrame([
        ["Certification-relevant denominator", "Only source × application-group records whose source-level Certification and Approval Basis is a certification, standard, or official approval category are included."],
        ["Included basis categories", "EN 13432 / OK compost / Seedling; Government/programme approval; BPI / ASTM / CMA; OK compost HOME / NF T 51-800 / AS 5810; Country-specific standards/certifications (BNQ / AS 4736 / DINplus)."],
        ["Excluded basis categories", "Generic compostable/biodegradable wording only and No named standard/approval stated are excluded from Certification Sufficiency, but remain reported under Certification and Approval Basis."],
    ], columns=["Item", "Description"])
    nrows, ncols = write_dataframe(ws, route3_note, row)
    style_range(ws, row, row + nrows - 1, 1, ncols, row)
    row += nrows + 2
    cert_basis_dist = shares[shares["indicator"] == "Certification and Approval Basis"].copy() if "indicator" in shares.columns else pd.DataFrame()
    if not cert_basis_dist.empty:
        row = add_section(ws, row, "3.1 Certification and Approval Basis", "What standards, labels, programme approvals, or generic compostability wording are mentioned in the source corpus?", 10)
        show = cert_basis_dist[[c for c in ["category", "n", "denominator", "share", "share_percent", "denominator_basis"] if c in cert_basis_dist.columns]].copy().sort_values(["n", "category"], ascending=[False, True])
        nrows, ncols = write_dataframe(ws, show, row)
        style_range(ws, row, row + nrows - 1, 1, ncols, row)
        format_numeric_columns(ws, show, row)
        row += nrows + 2

    if not cert_summary.empty:
        cert_summary = order_certification_sufficiency_summary(cert_summary)
        row = add_section(ws, row, "3.2 Certification Sufficiency", "Among certification-relevant records, does certification correspond to acceptance, conditional acceptance, rejection, or uncertainty?", 10)
        show = cert_summary[[c for c in ["certification_sufficiency_status", "n", "denominator", "share", "share_percent", "denominator_basis"] if c in cert_summary.columns]].copy()
        nrows, ncols = write_dataframe(ws, show, row)
        style_range(ws, row, row + nrows - 1, 1, ncols, row)
        format_numeric_columns(ws, show, row)
        row += nrows + 2

    if cert_by_app is not None and not cert_by_app.empty:
        row = add_section(ws, row, "3.3 Certification Sufficiency by Application Group", "Bubble-chart input: how do certification-sufficiency outcomes differ across the seven compostable-packaging application groups?", 10)
        app_cols = [
            "application_type", "evidence_records",
            "Certification accepted", "Certification conditionally accepted", "Certification rejected", "Certification unclear or not stated",
            "share__Certification accepted", "share__Certification conditionally accepted", "share__Certification rejected", "share__Certification unclear or not stated",
            "denominator_basis",
        ]
        show = cert_by_app[[c for c in app_cols if c in cert_by_app.columns]].copy()
        nrows, ncols = write_dataframe(ws, show, row)
        style_range(ws, row, row + nrows - 1, 1, ncols, row)
        format_numeric_columns(ws, show, row)
        row += nrows + 2

    # Cross-tabulations by source authority and certification basis are retained as supporting workbook sheets/CSV outputs.

    autofit_reasonable(ws, pd.DataFrame(columns=["category", "n", "denominator", "share", "share_percent", "denominator_basis", "x", "y", "z", "q"]))
    ws.freeze_panes = "A4"

def create_plain_table_sheet(wb: Workbook, sheet_name: str, title: str, subtitle: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    row = add_sheet_title(ws, title, subtitle, max(6, min(len(df.columns), 12)))
    nrows, ncols = write_dataframe(ws, df, row)
    style_range(ws, row, row + nrows - 1, 1, ncols, row)
    format_numeric_columns(ws, df, row)
    autofit_reasonable(ws, df)
    ws.freeze_panes = f"A{row+1}"


def create_certification_analysis_codebook_sheet(wb: Workbook, codebook: pd.DataFrame) -> None:
    indicators = [
        "Certification and Approval Basis",
        "Certification Sufficiency",
    ]
    sdf = codebook[codebook["indicator"].isin(indicators)].copy() if "indicator" in codebook.columns else pd.DataFrame()
    if sdf.empty:
        return
    # Keep certification-related and intermediate categories in one reader-facing guide.
    order_map = {name: i for i, name in enumerate(indicators)}
    sdf["_indicator_order"] = sdf["indicator"].map(order_map)
    cols = [
        "indicator", "category", "definition", "included_examples", "excluded_examples",
        "include_regex_examples", "exclude_regex_examples", "regex_or_rule_examples",
        "coding_basis", "priority_or_mapping_note",
    ]
    sdf = sdf.sort_values(["_indicator_order", "category"]).drop(columns=["_indicator_order"])
    sdf = sdf[[c for c in cols if c in sdf.columns]]
    create_plain_table_sheet(
        wb,
        "CB_Certification_Analysis",
        "Codebook: certification sufficiency analysis",
        "Guide for Certification and Approval Basis and the filtered Certification Sufficiency indicator. Certification Sufficiency uses only certification/standard/official-approval basis categories; generic wording and no-standard cases are excluded.",
        sdf,
    )


def create_codebook_sheets(wb: Workbook, codebook: pd.DataFrame) -> None:
    for indicator, sheet_name in CODEBOOK_SHEET_NAMES.items():
        sdf = codebook[codebook["indicator"] == indicator].copy()
        if sdf.empty:
            continue
        if indicator == "Rejection Rationale":
            cols = [
                "category",
                "main_share_denominator_status",
                "reporting_note",
                "definition",
                "included_examples",
                "excluded_examples",
                "include_regex_examples",
                "exclude_regex_examples",
                "regex_or_rule_examples",
                "coding_basis",
                "priority_or_mapping_note",
            ]
            subtitle = (
                "Definitions and reporting status for the rejection-rationale indicator. "
                "Only substantive stated rationale categories are included in the main share; "
                "Reason not explicitly stated and Not applicable remain audit statuses."
            )
        else:
            cols = ["category", "definition", "included_examples", "excluded_examples", "include_regex_examples", "exclude_regex_examples", "regex_or_rule_examples", "coding_basis", "priority_or_mapping_note"]
            subtitle = "Definitions and transparent regex/rule examples. For Source-level Sensitivity Group and Application-level Compatibility Scenario, mappings are shown instead of direct regex."
        sdf = sdf[[c for c in cols if c in sdf.columns]]
        create_plain_table_sheet(wb, sheet_name, f"Codebook: {indicator}", subtitle, sdf)


def create_certification_sufficiency_sheet(wb: Workbook, tables: Dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("06_Certification_Suff")
    ws.sheet_view.showGridLines = False
    row = add_sheet_title(
        ws,
        "Certification sufficiency",
        "Analysis route 3: certification sufficiency for source × application-group records whose source-level basis is a certification, standard, or official approval category.",
        10,
    )

    cert_basis_dist = tables.get("reference_shares", pd.DataFrame())
    if cert_basis_dist is not None and not cert_basis_dist.empty and "indicator" in cert_basis_dist.columns:
        cert_basis_dist = cert_basis_dist[cert_basis_dist["indicator"] == "Certification and Approval Basis"].copy()
    else:
        cert_basis_dist = pd.DataFrame()

    sections = [
        (
            "3.1 Certification and Approval Basis",
            "What standards, labels, programme approvals, or generic compostability wording are mentioned in the source corpus?",
            cert_basis_dist,
            ["category", "n", "denominator", "share", "share_percent", "denominator_basis"],
        ),
        (
            "3.2 Certification Sufficiency",
            "Among certification-relevant records, does certification correspond to acceptance, conditional acceptance, rejection, or uncertainty?",
            order_certification_sufficiency_summary(tables.get("certification_sufficiency_summary", pd.DataFrame())),
            ["certification_sufficiency_status", "n", "denominator", "share", "share_percent", "denominator_basis"],
        ),
        (
            "3.3 Certification Sufficiency by Application Group",
            "Wide bubble-chart input table with all seven application groups and all certification-sufficiency categories.",
            tables.get("certification_sufficiency_by_application", pd.DataFrame()),
            ["application_type", "evidence_records", "Certification accepted", "Certification conditionally accepted", "Certification rejected", "Certification unclear or not stated", "share__Certification accepted", "share__Certification conditionally accepted", "share__Certification rejected", "share__Certification unclear or not stated", "denominator_basis"],
        ),
        (
            "3.4 Certification Sufficiency Bubble Matrix",
            "Long-form bubble-chart input: one row per Application Group × Certification Sufficiency category.",
            tables.get("certification_sufficiency_bubble_matrix", pd.DataFrame()),
            ["application_type", "certification_sufficiency_status", "n", "application_denominator", "share_within_application", "denominator_basis"],
        ),
    ]

    for title, question, df, cols in sections:
        if df is None or df.empty:
            continue
        row = add_section(ws, row, title, question, 10)
        show = df[[c for c in cols if c in df.columns]].copy()
        nrows, ncols = write_dataframe(ws, show, row)
        style_range(ws, row, row + nrows - 1, 1, ncols, row)
        format_numeric_columns(ws, show, row)
        row += nrows + 2

    autofit_reasonable(ws, pd.DataFrame(columns=["application_type", "evidence_records", "Certification accepted", "Certification conditionally accepted", "Certification rejected", "Certification unclear or not stated", "mean_certification_sufficiency_score"]))
    ws.freeze_panes = "A4"


def citation_sort_value(value) -> int:
    m = re.search(r"\d+", str(value or ""))
    return int(m.group(0)) if m else 10**9


def sorted_reference_coding(ref: pd.DataFrame) -> pd.DataFrame:
    out = ref.copy()
    if "country_or_region" not in out.columns:
        return out
    out["_country_sort"] = out["country_or_region"].fillna("").astype(str).str.casefold()
    out["_citation_sort"] = out["citation_id"].map(citation_sort_value) if "citation_id" in out.columns else range(len(out))
    out = out.sort_values(["_country_sort", "_citation_sort"], kind="mergesort")
    return out.drop(columns=["_country_sort", "_citation_sort"])


def breakdown(series: pd.Series) -> str:
    values = series.fillna("Not coded").astype(str)
    counts = values.value_counts(dropna=False).sort_index()
    return "; ".join(f"{category}: {int(count)}" for category, count in counts.items())


def country_coverage_table(ref: pd.DataFrame) -> pd.DataFrame:
    required = {"country_or_region", "citation_id", "Source Authority", "Acceptance"}
    missing = required - set(ref.columns)
    if missing:
        raise ValueError(f"reference_coding missing columns for 05_Country_Coverage: {sorted(missing)}")

    rows = []
    work = ref.copy()
    work["country_or_region"] = work["country_or_region"].fillna("").astype(str)
    for country, sdf in work.groupby("country_or_region", dropna=False, sort=True):
        authorities = sdf["Source Authority"].fillna("").astype(str)
        national_mask = authorities.isin({"National rule or policy", "Supranational framework"})
        citation_ids = sorted((str(v) for v in sdf["citation_id"].tolist()), key=citation_sort_value)
        rows.append({
            "country_or_region": country,
            "n_sources_total": int(len(sdf)),
            "n_national_or_supranational": int(national_mask.sum()),
            "n_non_national_operational": int((~national_mask).sum()),
            "n_source_authority_types": int(authorities.replace("", pd.NA).dropna().nunique()),
            "source_authority_breakdown": breakdown(sdf["Source Authority"]),
            "acceptance_breakdown": breakdown(sdf["Acceptance"]),
            "citation_ids": "; ".join(citation_ids),
        })
    return pd.DataFrame(rows, columns=[
        "country_or_region",
        "n_sources_total",
        "n_national_or_supranational",
        "n_non_national_operational",
        "n_source_authority_types",
        "source_authority_breakdown",
        "acceptance_breakdown",
        "citation_ids",
    ])




def application_by_country_table(ref: pd.DataFrame, app: pd.DataFrame) -> pd.DataFrame:
    """Create a country-grouped application-level view for manuscript support.

    The first five columns are manuscript-ready. They combine source-level
    coverage and application-level coverage in a compact "Evidence coverage"
    column so the table can be copied directly into the manuscript.

    Compatible application groups follow the central scenario definition:
    Accepted + Accepted with conditions. This is a descriptive country-grouped
    view, not a nationally representative country estimate.
    """
    ref_required = {"citation_id", "country_or_region", "subnational_or_context", "Acceptance"}
    app_required = {"citation_id", "country_or_region", "application_type", "application_decision"}
    ref_missing = ref_required - set(ref.columns)
    app_missing = app_required - set(app.columns)
    if ref_missing:
        raise ValueError(f"reference_coding missing columns for 05_Application_by_Country: {sorted(ref_missing)}")
    if app_missing:
        raise ValueError(f"application_long missing columns for 05_Application_by_Country: {sorted(app_missing)}")

    ref_work = ref.copy()
    app_work = app.copy()
    ref_work["citation_id"] = ref_work["citation_id"].astype(str)
    app_work["citation_id"] = app_work["citation_id"].astype(str)
    ref_work["country_or_region"] = ref_work["country_or_region"].fillna("").astype(str)
    app_work["country_or_region"] = app_work["country_or_region"].fillna("").astype(str)

    compatible_decisions = {"Accepted", "Accepted with conditions"}

    def unique_join(values: pd.Series) -> str:
        vals = [str(v).strip() for v in values.fillna("").tolist()]
        vals = sorted({v for v in vals if v}, key=str.casefold)
        return "; ".join(vals)

    def application_group_summary(adf: pd.DataFrame, decision_set: set[str], status_word: str) -> str:
        """Return application groups with counts as 'Application (n/total records status)'.

        The denominator is all source × application-group records for the same
        country and Application Group, not all sources in that country.
        """
        if adf.empty:
            return "No application-level records generated; see source-level Acceptance"
        parts = []
        for app_type, gdf in adf.groupby("application_type", dropna=False, sort=True):
            app_name = str(app_type).strip()
            if not app_name or app_name.lower() == "nan":
                continue
            decisions = gdf["application_decision"].fillna("Unclear or not stated").astype(str)
            n_status = int(decisions.isin(decision_set).sum())
            n_total = int(len(gdf))
            if n_status > 0:
                parts.append((app_name.casefold(), f"{app_name} ({n_status}/{n_total} records {status_word})"))
        if not parts:
            if status_word == "compatible":
                return "No compatible application group observed"
            if status_word == "rejected":
                return "No rejected application group observed"
            return "None observed"
        return "; ".join(text for _, text in sorted(parts, key=lambda x: x[0]))

    rows = []
    all_countries = sorted(
        set(ref_work["country_or_region"].dropna().astype(str)) | set(app_work["country_or_region"].dropna().astype(str)),
        key=str.casefold,
    )
    for country in all_countries:
        sdf = ref_work[ref_work["country_or_region"].eq(country)]
        adf = app_work[app_work["country_or_region"].eq(country)]

        source_citation_ids = sorted((str(v) for v in sdf.get("citation_id", pd.Series(dtype=object)).tolist()), key=citation_sort_value)
        app_citation_ids = sorted((str(v) for v in adf.get("citation_id", pd.Series(dtype=object)).unique().tolist()), key=citation_sort_value)

        n_sources = int(len(sdf))
        n_app_records = int(len(adf))
        evidence_coverage = f"{n_sources} sources; {n_app_records} source × application-group records"

        if n_app_records == 0:
            coverage_note = "Source-level evidence only; no application-level records generated"
        elif n_sources <= 1:
            coverage_note = "Single-source country observation; not nationally representative"
        elif n_sources <= 4:
            coverage_note = "Limited observed source coverage; not nationally representative"
        else:
            coverage_note = "Multiple observed sources; still not nationally representative"

        rows.append({
            # Manuscript-ready front columns
            "Country / region": country,
            "Evidence coverage": evidence_coverage,
            "Covered jurisdictions / contexts": unique_join(sdf["subnational_or_context"]) if not sdf.empty else "",
            "Observed compatible Application Group categories": application_group_summary(adf, compatible_decisions, "compatible"),
            "Observed rejected Application Group categories": application_group_summary(adf, {"Rejected"}, "rejected"),
            # Diagnostic/supporting columns for sorting, checking, and transparency
            "n sources": n_sources,
            "n source × application-group records": n_app_records,
            "Source-level Acceptance breakdown": breakdown(sdf["Acceptance"]) if not sdf.empty else "No source-level records",
            "Application Decision breakdown": breakdown(adf["application_decision"]) if not adf.empty else "No application-level records generated",
            "Coverage note": coverage_note,
            "Source citation IDs": "; ".join(source_citation_ids),
            "Application-record citation IDs": "; ".join(app_citation_ids),
            "Interpretation note": "Country-grouped descriptive view only; source coverage is uneven and not nationally representative. Compatible groups require at least one source × application-group record coded Accepted or Accepted with conditions under the central scenario. The same Application Group may appear in both compatible and rejected columns when different retained sources within the same country give different application-level decisions.",
        })

    columns = [
        # Manuscript-ready columns
        "Country / region",
        "Evidence coverage",
        "Covered jurisdictions / contexts",
        "Observed compatible Application Group categories",
        "Observed rejected Application Group categories",
        # Diagnostic/supporting columns
        "n sources",
        "n source × application-group records",
        "Source-level Acceptance breakdown",
        "Application Decision breakdown",
        "Coverage note",
        "Source citation IDs",
        "Application-record citation IDs",
        "Interpretation note",
    ]
    return (
        pd.DataFrame(rows, columns=columns)
        .sort_values(["n sources", "Country / region"], ascending=[False, True], kind="mergesort")
        .reset_index(drop=True)
    )

def create_application_compatibility_codebook_sheet(wb: Workbook, codebook: pd.DataFrame, compat: pd.DataFrame) -> None:
    rows = [
        ["Concept", "Application-level Compatibility Scenario", "A tabulated analytical output, not a simple categorical indicator. It combines Application Group, Application Decision, and scenario assumptions to calculate compatibility shares."],
        ["Scenario", "Lower-bound scenario", "Counts only accepted records as compatible. Accepted-with-conditions and rejected records are incompatible. Unclear/not-stated records are excluded from the denominator."],
        ["Scenario", "Central scenario", "Counts accepted and accepted-with-conditions records as compatible. Rejected records are incompatible. Unclear/not-stated records are excluded from the denominator."],
        ["Scenario", "Upper-bound scenario", "Counts accepted and accepted-with-conditions records as compatible and treats unclear/not-stated records as potentially compatible. Rejected records remain incompatible."],
        ["Coding input", "Application Group", "Defined in CB_Application_Group."],
        ["Coding input", "Application Decision", "Defined in CB_Application_Decision."],
    ]
    df = pd.DataFrame(rows, columns=["Item type", "Item", "Definition / use"])
    create_plain_table_sheet(
        wb,
        "CB_Application_Scenario",
        "Codebook: Application-level Compatibility Scenario",
        "This sheet defines the scenario conversion table. The two coded dimensions used to construct it are documented separately in CB_Application_Group and CB_Application_Decision.",
        df,
    )


def source_level_display_columns(ref: pd.DataFrame) -> pd.DataFrame:
    """Return only metadata plus agreed source-level indicators for the reader-facing sheet."""
    cols = [
        "citation_id",
        "country_or_region",
        "subnational_or_context",
        "source_label",
        "source_title",
        "source_url",
        "Source Authority",
        "Stated Organic-waste Treatment Route",
        "Acceptance",
        "Rejection Rationale",
        "Certification and Approval Basis",
        "Source-level Sensitivity Group",
    ]
    return ref[[c for c in cols if c in ref.columns]].copy()


def application_records_display_columns(app: pd.DataFrame) -> pd.DataFrame:
    """Return the two coded application dimensions plus minimal source metadata."""
    cols = [
        "citation_id",
        "country_or_region",
        "source_label",
        "source_title",
        "source_url",
        "application_type",
        "application_decision",
        "decision_basis",
    ]
    out = app[[c for c in cols if c in app.columns]].copy()
    return out.rename(columns={
        "application_type": "Application Group",
        "application_decision": "Application Decision",
        "decision_basis": "Decision Basis",
    })

# -----------------------------------------------------------------------------
# Final workbook layout layer
# -----------------------------------------------------------------------------

# The layout functions below intentionally avoid sheet-specific hardcoded column
# widths.  They infer widths from the actual cell contents and from semantic
# column/header names, then cap widths/heights so the workbook stays readable
# when categories or wording change later.

SHORT_WIDTH = (9, 18)
MEDIUM_WIDTH = (16, 34)
LONG_WIDTH = (28, 72)
URL_WIDTH = (32, 64)


def _semantic_width_limits(header: str) -> tuple[float, float]:
    """Return width limits from a column/header name, not from a sheet name."""
    h = (header or "").strip().lower()
    if any(k in h for k in ["url", "link"]):
        return URL_WIDTH
    if any(k in h for k in [
        "definition", "description", "interpretation", "rationale", "basis",
        "note", "examples", "source title", "source label", "decision basis",
        "denominator basis", "included", "excluded", "regex", "rule", "question",
        "breakdown", "source wording",
    ]):
        return LONG_WIDTH
    if any(k in h for k in [
        "authority", "route", "acceptance", "certification", "application",
        "decision", "category", "indicator", "country", "region", "municipality",
        "source organization", "sensitivity",
    ]):
        return MEDIUM_WIDTH
    if any(k in h for k in [
        "n", "value", "share", "percent", "denominator", "id", "source id",
        "priority", "count", "records",
    ]):
        return SHORT_WIDTH
    return MEDIUM_WIDTH


def _visible_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def _quantile(values: list[int], q: float, default: int = 10) -> int:
    vals = sorted(v for v in values if v is not None)
    if not vals:
        return default
    idx = min(len(vals) - 1, max(0, int(round(q * (len(vals) - 1)))))
    return vals[idx]


def _infer_column_width(ws, col_idx: int, sample_rows: int = 250) -> float:
    """Infer a balanced width from content, using q90 not max to avoid one outlier."""
    header = _visible_text(ws.cell(1, col_idx).value)
    min_w, max_w = _semantic_width_limits(header)
    lengths: list[int] = [len(header)]
    max_row = min(ws.max_row, sample_rows)
    for row_idx in range(1, max_row + 1):
        text = _visible_text(ws.cell(row_idx, col_idx).value)
        if not text:
            continue
        # Treat line breaks as independent line lengths.
        lengths.extend(len(part) for part in text.splitlines() if part)
    q90 = _quantile(lengths, 0.90, default=len(header) or 10)
    q98 = _quantile(lengths, 0.98, default=q90)
    # Blend q90 and q98: responsive to long content but not dominated by URLs or outliers.
    estimated = max(len(header) + 2, q90 * 0.90 + q98 * 0.20 + 2)
    return round(min(max(estimated, min_w), max_w), 1)


def _infer_row_height(ws, row_idx: int, max_col: int, *, min_h: float = 18, max_h: float = 90) -> float:
    """Estimate row height from current inferred column widths and wrapped text."""
    max_lines = 1
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row_idx, col_idx)
        text = _visible_text(cell.value)
        if not text:
            continue
        width = ws.column_dimensions[get_column_letter(col_idx)].width or 18
        # Excel width roughly corresponds to characters.  Use a conservative factor
        # because Calibri 11/12 wraps earlier than monospace estimates.
        chars_per_line = max(int(width * 1.15), 10)
        manual_parts = text.splitlines() or [text]
        needed = 0
        for part in manual_parts:
            needed += max(1, int((len(part) + chars_per_line - 1) / chars_per_line))
        max_lines = max(max_lines, needed)
    return min(max(min_h, 15 * max_lines + 4), max_h)


def apply_adaptive_workbook_layout(wb: Workbook) -> None:
    """Apply content-adaptive layout to all reader-facing workbook sheets.

    The algorithm is deliberately generic:
    1. infer column-width limits from column/header semantics,
    2. use actual content lengths to choose widths,
    3. estimate row heights after widths are known,
    4. cap both widths and heights to avoid unreadable extremes.
    """
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        max_col = max(ws.max_column, 1)
        max_row = max(ws.max_row, 1)

        # Use the first non-empty row as the header proxy for plain table/codebook sheets.
        # Merged title rows are still handled acceptably because the content columns below
        # drive the q90/q98 width estimate.
        for col_idx in range(1, max_col + 1):
            width = _infer_column_width(ws, col_idx)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for row_idx in range(1, min(max_row, 1200) + 1):
            if row_idx == 1:
                ws.row_dimensions[row_idx].height = 30
            elif row_idx == 2:
                ws.row_dimensions[row_idx].height = min(max(_infer_row_height(ws, row_idx, max_col, min_h=26, max_h=56), 34), 56)
            else:
                ws.row_dimensions[row_idx].height = _infer_row_height(ws, row_idx, max_col, min_h=20, max_h=86)

        # Freeze panes generically for sheets with title/subtitle + table header.
        if ws.max_row >= 5 and not ws.title.startswith("00_Read_Me"):
            ws.freeze_panes = "A5"

        # A few visual safety tweaks based on structure, not exact widths.
        if ws.title == "00_Result_Summary":
            # Summary rows often contain section/question text and benefit from a little more space.
            for r in range(1, min(ws.max_row, 140) + 1):
                val = _visible_text(ws.cell(r, 1).value)
                if val.startswith(("0.", "1.", "2.", "3.")) or val.startswith("Analysis route"):
                    ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 24, 28)
        elif ws.title.startswith("CB_"):
            # Codebooks carry regex examples, so allow a slightly higher cap.
            for r in range(5, min(ws.max_row, 160) + 1):
                ws.row_dimensions[r].height = min(max(ws.row_dimensions[r].height or 24, 24), 110)


def build_workbook(tables: Dict[str, pd.DataFrame], source_desc: str) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    ref_sorted = sorted_reference_coding(tables["reference_coding"])
    ref_sorted = ref_sorted.drop(columns=["primary_source_for_sensitivity"], errors="ignore")
    source_display = source_level_display_columns(ref_sorted)
    app_display = application_records_display_columns(tables["application_long"])

    create_readme(wb, source_desc)
    create_result_summary(wb, tables)
    create_plain_table_sheet(
        wb,
        "01_Source_Level_Coding",
        "Clean source-level coding",
        "Source metadata followed by the agreed source-level analytical indicators. Treatment-route audit notes and regex traces are kept outside this reader-facing sheet.",
        source_display,
    )
    create_plain_table_sheet(
        wb,
        "02_Application_Records",
        "Application-level records",
        "One row per source × application group. Application Group and Application Decision are the two coded dimensions used to construct the application-level compatibility scenario output in 00_Result_Summary.",
        app_display,
    )
    create_plain_table_sheet(wb, "03_Source_Level_Sensitivity", "Source-level sensitivity", "Source-level indicator shares recalculated for all included sources, national/supranational sources, and sources below national level.", tables["source_level_sensitivity"])
    create_plain_table_sheet(
        wb,
        "04_Country_Coverage",
        "Country/region source coverage",
        "Factual descriptive counts by country or region. No prioritization or new categorical indicators are added.",
        country_coverage_table(ref_sorted),
    )
    create_plain_table_sheet(
        wb,
        "05_Application_by_Country",
        "Application-level compatibility by country/region",
        "Columns A–E are manuscript-ready. Compatible application groups follow the central scenario: at least one source × application-group record coded Accepted or Accepted with conditions. Rejected groups require at least one source × application-group record coded Rejected. Country grouping is descriptive and not nationally representative.",
        application_by_country_table(ref_sorted, tables["application_long"]),
    )
    create_certification_sufficiency_sheet(wb, tables)
    create_application_compatibility_codebook_sheet(wb, tables.get("codebook", pd.DataFrame()), tables.get("application_compatibility", pd.DataFrame()))
    create_codebook_sheets(wb, tables["codebook"])
    apply_adaptive_workbook_layout(wb)
    return wb


def validate_required_columns(tables: Dict[str, pd.DataFrame]) -> None:
    checks = {
        "reference_coding": ["citation_id", "Source Authority", "Acceptance", "Source-level Sensitivity Group"],
        "application_long": ["citation_id", "application_type", "application_decision", "decision_basis"],
        "application_compatibility": ["application_type", "evidence_records", "central_compatible_share"],
        "certification_sufficiency_summary": ["certification_sufficiency_status", "n", "denominator", "share"],
        "certification_sufficiency_by_application": ["application_type", "evidence_records", "Certification accepted", "Certification rejected"],
        "certification_sufficiency_bubble_matrix": ["application_type", "certification_sufficiency_status", "n", "application_denominator"],
        "certification_sufficiency_by_certification_basis": ["Certification and Approval Basis", "evidence_records", "Accepted", "Rejected"],
        "certification_sufficiency_by_source_authority": ["Source Authority", "evidence_records"],
        "certification_sufficiency_long": ["citation_id", "application_type", "application_decision", "certification_sufficiency_status", "certification_sufficiency_score"],
        "source_level_sensitivity": ["indicator", "source_level_group", "category", "n", "denominator", "share"],
        "codebook": ["indicator", "category", "definition", "include_regex_examples", "exclude_regex_examples", "regex_or_rule_examples", "coding_basis"],
    }
    errors = []
    for key, cols in checks.items():
        missing = [c for c in cols if c not in tables[key].columns]
        if missing:
            errors.append(f"{REQUIRED_FILES[key]} missing columns: {missing}")
    if errors:
        raise ValueError("\n".join(errors))


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate the Supplementary Data Excel workbook.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="Path to results folder, display folder, or display zip")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT_XLSX), help="Output .xlsx path")
    args = parser.parse_args()

    input_path = resolve_path(args.input)
    output = resolve_path(args.output)
    tables, source_desc = read_display_tables(input_path)
    tables = align_table_terminology(tables)
    tables = augment_rejection_rationale_codebook(tables)
    validate_required_columns(tables)
    tables = {k: harmonize_certification_sufficiency_labels(harmonize_certification_category_labels(v)) for k, v in tables.items()}
    wb = build_workbook(tables, source_desc)
    wb.save(output)

    print(f"Workbook written: {output}")
    print(f"Sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()